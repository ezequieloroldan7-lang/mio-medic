"""
migrate.py — Importa pacientes del Excel a la base de datos.
Uso desde la carpeta /backend:
    python migrate.py ../pacientes.xlsx
    python migrate.py ../pacientes.xlsx --dry-run     # no toca la BD, solo muestra
    python migrate.py ../pacientes.xlsx --reset       # borra pacientes previos

Sin pandas: usa openpyxl directamente, así funciona con Python 3.14.
"""
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from database import SessionLocal, engine
import models


# ── Helpers ──────────────────────────────────────────────────
def _s(v) -> str:
    """Valor como string limpio, o cadena vacía."""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none") else s


def limpiar_telefono(v) -> str | None:
    s = _s(v)
    if not s:
        return None
    # Viene como float (1168708761.0) → quitamos el .0
    if re.match(r"^-?\d+\.0+$", s):
        s = s.split(".", 1)[0]
    s = re.sub(r"[\s\-()]", "", s)
    if not s:
        return None
    if s.startswith("+"):
        return s
    if s.startswith("54"):
        return "+" + s
    if s.startswith("0"):
        return "+54" + s[1:]
    return "+54" + s


def normalizar_cobertura(v) -> str | None:
    s = _s(v)
    if not s:
        return None
    low = s.lower().replace(" ", "")
    mapa = {
        "osde210": "OSDE 210",
        "osde310": "OSDE 310",
        "osde410": "OSDE 410",
        "osde450": "OSDE 450",
        "osde/privada": "OSDE / Particular",
        "osde": "OSDE",
        "medicus": "Medicus",
        "medife": "Medife",
        "smg": "Swiss Medical",
        "s.medical": "Swiss Medical",
        "swissmedical": "Swiss Medical",
        "omint": "Omint",
        "obsba": "OBSBA",
        "ospe": "OSPE",
        "poderjudicial": "Poder Judicial",
        "privada": "Particular",
        "particular": "Particular",
    }
    return mapa.get(low, s.title())


def _hc(v) -> str | None:
    s = _s(v)
    if not s:
        return None
    if re.match(r"^-?\d+\.0+$", s):
        return s.split(".", 1)[0]
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s


def _split_nombre(completo: str) -> tuple[str, str]:
    """
    El Excel tiene formato 'Apellido Nombre Segundo...'.
    Convención: primera palabra = apellido, resto = nombre.
    """
    partes = completo.split()
    if not partes:
        return "", ""
    if len(partes) == 1:
        return partes[0], ""
    return partes[0], " ".join(partes[1:])


# ── Importación principal ────────────────────────────────────
HEADERS_MAP = {
    "paciente":    "paciente",
    "telefono":    "telefono",
    "teléfono":    "telefono",
    "email":       "email",
    "deriva":      "deriva",
    "cobertura":   "cobertura",
    "hc":          "hc",
    "miomedic":    "miomedic",
    "se atiende con": "deriva_extra",
}


def migrar(path_excel: str, dry_run: bool = False, reset: bool = False):
    ruta = Path(path_excel).resolve()
    if not ruta.exists():
        print(f"[ERROR] No existe el archivo: {ruta}")
        sys.exit(1)

    print(f"[INFO] Leyendo {ruta} ...")
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    # Mapear índices de columnas por nombre
    col_idx = {}
    for i, h in enumerate(header):
        key = _s(h).lower()
        if key in HEADERS_MAP:
            col_idx[HEADERS_MAP[key]] = i

    if "paciente" not in col_idx:
        print("[ERROR] El Excel no tiene columna 'Paciente'. Headers:", header)
        sys.exit(1)

    print(f"   Columnas detectadas: {list(col_idx.keys())}")

    models.Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    try:
        if reset and not dry_run:
            borrados = db.query(models.Paciente).delete()
            db.commit()
            print(f"[RESET] {borrados} pacientes previos eliminados.")

        # Índice de pacientes existentes para evitar duplicados (por apellido+nombre)
        existentes = set()
        if not reset:
            for ap, nm in db.query(models.Paciente.apellido, models.Paciente.nombre).all():
                existentes.add((ap.strip().lower(), (nm or "").strip().lower()))

        insertados = 0
        duplicados = 0
        omitidos = 0
        muestras = []

        for row in rows:
            if not row or all(c is None for c in row):
                continue
            nombre_completo = _s(row[col_idx["paciente"]])
            if not nombre_completo:
                omitidos += 1
                continue

            apellido, nombre = _split_nombre(nombre_completo)
            key = (apellido.lower(), nombre.lower())
            if key in existentes:
                duplicados += 1
                continue

            p = models.Paciente(
                nombre    = nombre,
                apellido  = apellido,
                telefono  = limpiar_telefono(row[col_idx["telefono"]]) if "telefono" in col_idx else None,
                email     = _s(row[col_idx["email"]]) or None if "email" in col_idx else None,
                nro_hc    = _hc(row[col_idx["hc"]]) if "hc" in col_idx else None,
                financiador = normalizar_cobertura(row[col_idx["cobertura"]]).upper() if "cobertura" in col_idx and normalizar_cobertura(row[col_idx["cobertura"]]) else None,
                deriva    = _s(row[col_idx["deriva"]]) or None if "deriva" in col_idx else None,
            )

            if dry_run:
                if len(muestras) < 8:
                    muestras.append(
                        f"{p.apellido}, {p.nombre} | tel={p.telefono} | hc={p.nro_hc} | fin={p.financiador}"
                    )
            else:
                db.add(p)

            existentes.add(key)
            insertados += 1

            # Commit por lotes para evitar transacciones enormes
            if not dry_run and insertados % 200 == 0:
                db.commit()

        if not dry_run:
            db.commit()

        print("")
        print("[OK] Importacion terminada.")
        print(f"     - Insertados : {insertados}")
        print(f"     - Duplicados : {duplicados}")
        print(f"     - Omitidos   : {omitidos}")
        if dry_run:
            print("\n[DRY-RUN] Primeras filas que se insertarian:")
            for m in muestras:
                print("   -", m)
            print("\n(dry-run: no se modifico la BD)")
    finally:
        db.close()


# ── CLI ──────────────────────────────────────────────────────
if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] in ("-h", "--help"):
        print("Uso: python migrate.py <ruta_al_excel> [--dry-run] [--reset]")
        sys.exit(0)
    path = args[0]
    dry_run = "--dry-run" in args
    reset   = "--reset"   in args
    migrar(path, dry_run=dry_run, reset=reset)
